"""
Excel report generation via openpyxl.

v4.0: Styled Excel spreadsheets for matches, ranking, and players data.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# Color palette
ACCENT_FILL = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0A0E17", end_color="0A0E17", fill_type="solid")
STRIPE_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
HEADER_FONT = Font(color="F8FAFC", bold=True, size=10)
ACCENT_FONT = Font(color="00D4AA", bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


async def generate_excel_report(
    client: Any, data_type: str, page: int = 1,
) -> bytes:
    """Generate a styled Excel spreadsheet from HLTV data.

    Args:
        client: HLTVClient instance.
        data_type: 'matches', 'ranking', or 'players'.
        page: Page number for paginated data.

    Returns:
        Excel file as bytes (.xlsx).
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.title = data_type.title()

    # ── Title row ─────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"HLTV Pro v4.0 — {data_type.title()} Report"
    title_cell.font = ACCENT_FONT
    title_cell.alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:E2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {now} | Data: HLTV.org"
    sub_cell.font = Font(color="94A3B8", size=9)

    if data_type == "ranking":
        from src.endpoints.teams import TeamsEndpoint

        ranking_data = await TeamsEndpoint(client).get_ranking()
        teams = ranking_data.teams or []

        headers = ["Rank", "Team", "Points", "Change"]
        _write_header_row(ws, headers, row=4)

        for i, team in enumerate(teams[:100]):
            row = 5 + i
            ws.cell(row=row, column=1, value=team.rank)
            ws.cell(row=row, column=2, value=team.name)
            ws.cell(row=row, column=3, value=team.points)
            change = team.change
            ws.cell(
                row=row, column=4,
                value=f"+{change}" if change and change > 0 else (change or 0),
            )
            _style_data_row(ws, row, cols=4, stripe=(i % 2 == 1))

        _auto_width(ws, [8, 40, 10, 10])

    elif data_type == "matches":
        from src.endpoints.matches import MatchesEndpoint

        matches_data = await MatchesEndpoint(client).get_results(page=page)

        headers = ["Date", "Team 1", "Score", "Team 2", "Event"]
        _write_header_row(ws, headers, row=4)

        for i, match in enumerate(matches_data or []):
            row = 5 + i
            date_str = (
                match.date.strftime("%Y-%m-%d")
                if hasattr(match, "date") and match.date
                else "-"
            )
            s1 = getattr(match.team1, "score", None)
            s2 = getattr(match.team2, "score", None)
            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=match.team1.name)
            ws.cell(row=row, column=3, value=f"{s1 or '-'}")
            ws.cell(row=row, column=4, value=match.team2.name)
            ws.cell(row=row, column=5, value=f"{s2 or '-'}")
            evt_col = 6
            evt_name = (
                match.event.name
                if hasattr(match, "event") and match.event
                else "-"
            )
            ws.cell(row=row, column=evt_col, value=evt_name)
            _style_data_row(ws, row, cols=6, stripe=(i % 2 == 1))

        _auto_width(ws, [12, 25, 8, 25, 8, 25])

    elif data_type == "players":
        from src.endpoints.players import PlayersEndpoint

        players_data = await PlayersEndpoint(client).get_top_players("last3months")
        players = players_data.players if hasattr(players_data, "players") else []

        headers = ["Rank", "Player", "Team", "Rating", "Maps Played"]
        _write_header_row(ws, headers, row=4)

        for i, p in enumerate(players[:100]):
            row = 5 + i
            player_name = getattr(p.player, "name", "—") if p.player else "—"
            team_name = getattr(p.team, "name", "—") if p.team else "—"
            ws.cell(row=row, column=1, value=p.rank)
            ws.cell(row=row, column=2, value=player_name)
            ws.cell(row=row, column=3, value=team_name)
            ws.cell(row=row, column=4, value=round(p.rating, 2))
            ws.cell(row=row, column=5, value=p.maps_played)
            _style_data_row(ws, row, cols=5, stripe=(i % 2 == 1))

        _auto_width(ws, [8, 30, 25, 10, 14])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_header_row(ws, headers: list[str], row: int = 1) -> None:
    """Write and style a header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, cols: int, stripe: bool = False) -> None:
    """Apply border and optional stripe to a data row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if stripe:
            cell.fill = STRIPE_FILL


def _auto_width(ws, col_widths: list[int]) -> None:
    """Set approximate column widths."""
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
